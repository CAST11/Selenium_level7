import openpyxl

def get_row_count(file, sheet_name):
    """
    Returns total number of rows in the Excel sheet.
    """
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(file, sheet_name):
    """
    Returns total number of columns in the Excel sheet.
    """
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(file, sheet_name, row_num, col_num):
    """
    Reads a single cell value.
    """
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num, column=col_num).value


def write_data(file, sheet_name, row_num, col_num, data):
    """
    Writes data into a cell and saves the file.
    """
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=col_num).value = data
    workbook.save(file)